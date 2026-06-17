#!/usr/bin/env python3
"""
Extrae un resumen diario desde el Libro de Caja de Sterafarma.

Uso sencillo desde esta carpeta:
  python extraer_libro_caja.py

Uso indicando archivo:
  python extraer_libro_caja.py libro_de_caja20260617171340.xlsx
  python extraer_libro_caja.py libro_de_caja20260617171340.xlsx -o resumen.csv

Si no indicas archivo, busca el Excel libro_de_caja*.xlsx mas reciente en esta carpeta.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala dependencias con: python -m pip install -r requirements.txt")

FORMAS_MARKER = "DESGLOSE POR FORMAS DE PAGO"
HEADER_MARKER = "FORMA PAGO"
OUTPUT_COLUMNS = [
    "fecha",
    "dia_semana",
    "ingresos_tarjeta",
    "ingresos_efectivo",
    "salidas_caja",
    "retiradas_efectivo",
    "entradas_efectivo",
    "total_tarjeta",
    "total_efectivo",
]
WEEKDAYS_ES = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrae ingresos de tarjeta, efectivo, salidas y retiradas desde Sterafarma."
    )
    parser.add_argument(
        "input",
        nargs="?",
        help="Excel de entrada. Si se omite, usa el libro_de_caja*.xlsx mas reciente.",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Salida .xlsx o .csv. Si se omite, crea resumen_<nombre_input>.xlsx.",
    )
    parser.add_argument(
        "--date-mode",
        choices=["auto", "calendar", "non-sunday"],
        default="auto",
        help="auto, calendar o non-sunday. Por defecto intenta cuadrar por dias no domingo.",
    )
    return parser.parse_args()


def resolve_input(input_arg: str | None, base_dir: Path) -> Path:
    if input_arg:
        path = Path(input_arg)
        if not path.is_absolute():
            path = base_dir / path
        return path.resolve()

    candidates = sorted(
        base_dir.glob("libro_de_caja*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No encuentro ningun libro_de_caja*.xlsx en esta carpeta.")
    return candidates[0].resolve()


def as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def parse_excel_datetime(value: Any) -> dt.datetime:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"No he podido leer la fecha: {value!r}")


def normalize_label(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def read_date_range(ws: Any) -> tuple[dt.date, dt.date]:
    start = None
    end = None
    for row in ws.iter_rows():
        label = normalize_label(row[0].value if row else None)
        if label.startswith("FECHA INICIO"):
            start = parse_excel_datetime(row[1].value).date()
        elif label.startswith("FECHA FIN"):
            end = parse_excel_datetime(row[1].value).date()
        if start and end:
            return start, end
    raise ValueError("No encuentro 'Fecha inicio' y 'Fecha fin' en el Excel.")


def build_dates(start: dt.date, end: dt.date, mode: str) -> list[dt.date]:
    days = [start + dt.timedelta(days=i) for i in range((end - start).days + 1)]
    if mode == "non-sunday":
        return [day for day in days if day.weekday() != 6]
    return days


def assign_dates(block_count: int, start: dt.date, end: dt.date, mode: str) -> tuple[list[dt.date | None], str]:
    if mode != "auto":
        dates = build_dates(start, end, mode)
        warning = "" if len(dates) == block_count else (
            f"Hay {block_count} bloques, pero el modo {mode} genera {len(dates)} fechas."
        )
        return (dates + [None] * block_count)[:block_count], warning

    calendar_days = build_dates(start, end, "calendar")
    non_sundays = build_dates(start, end, "non-sunday")
    if len(calendar_days) == block_count:
        return calendar_days, ""
    if len(non_sundays) == block_count:
        return non_sundays, ""

    warning = (
        f"No cuadran las fechas: {block_count} bloques, "
        f"{len(calendar_days)} dias naturales y {len(non_sundays)} dias no domingo."
    )
    return (calendar_days + [None] * block_count)[:block_count], warning


def find_payment_blocks(ws: Any) -> list[int]:
    return [row for row in range(1, ws.max_row + 1) if normalize_label(ws.cell(row, 1).value) == FORMAS_MARKER]


def extract_payment_block(ws: Any, marker_row: int) -> dict[str, Any]:
    header_row = marker_row + 1
    if normalize_label(ws.cell(header_row, 1).value) != HEADER_MARKER:
        raise ValueError(f"Formato inesperado cerca de la fila {marker_row}.")

    payments: dict[str, dict[str, float]] = {}
    row = header_row + 1
    while row <= ws.max_row:
        label = normalize_label(ws.cell(row, 1).value)
        if not label or label == "TOTAL" or label.startswith("DESGLOSE "):
            break
        payments[label] = {
            "ventas_cob_credito": as_number(ws.cell(row, 2).value),
            "entradas_cta": as_number(ws.cell(row, 3).value),
            "entradas": as_number(ws.cell(row, 4).value),
            "salidas": as_number(ws.cell(row, 5).value),
            "total": as_number(ws.cell(row, 6).value),
            "ret_arqueo": as_number(ws.cell(row, 7).value),
        }
        row += 1

    tarjeta = payments.get("TARJETA", {})
    efectivo = payments.get("EFECTIVO", {})
    otros = {
        name: values["total"]
        for name, values in payments.items()
        if name not in {"TARJETA", "EFECTIVO"}
    }

    return {
        "ingresos_tarjeta": tarjeta.get("ventas_cob_credito", 0.0),
        "ingresos_efectivo": efectivo.get("ventas_cob_credito", 0.0),
        "salidas_caja": efectivo.get("salidas", 0.0),
        "retiradas_efectivo": efectivo.get("ret_arqueo", 0.0),
        "entradas_efectivo": efectivo.get("entradas", 0.0),
        "total_tarjeta": tarjeta.get("total", 0.0),
        "total_efectivo": efectivo.get("total", 0.0),
        "otros_medios_pago": "; ".join(f"{k}: {v:.2f}" for k, v in sorted(otros.items())),
        "fila_origen": marker_row,
    }


def extract_cashbook(input_path: Path, date_mode: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    start, end = read_date_range(ws)
    block_rows = find_payment_blocks(ws)
    dates, warning = assign_dates(len(block_rows), start, end, date_mode)

    records = []
    for index, marker_row in enumerate(block_rows):
        record = extract_payment_block(ws, marker_row)
        date = dates[index]
        record["fecha"] = date.isoformat() if date else ""
        record["dia_semana"] = WEEKDAYS_ES[date.weekday()] if date else ""
        record["aviso_fecha"] = warning
        records.append(record)
    return records


def format_decimal(value: Any) -> Any:
    if isinstance(value, float):
        return f"{value:.2f}".replace(".", ",")
    return value


def write_csv(records: list[dict[str, Any]], output_path: Path) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS, delimiter=";")
        writer.writeheader()
        for record in records:
            writer.writerow({col: format_decimal(record.get(col, "")) for col in OUTPUT_COLUMNS})


def write_xlsx(records: list[dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen diario"
    ws.append(OUTPUT_COLUMNS)
    for record in records:
        ws.append([record.get(col, "") for col in OUTPUT_COLUMNS])

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    money_cols = range(3, 10)
    for col in money_cols:
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).number_format = '#,##0.00'

    for col_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width, 12), 34)

    wb.save(output_path)


def default_output(input_path: Path) -> Path:
    return input_path.with_name(f"resumen_{input_path.stem}.xlsx")


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    input_path = resolve_input(args.input, base_dir)
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_path = Path(args.output).resolve() if args.output else default_output(input_path)
    records = extract_cashbook(input_path, args.date_mode)
    if not records:
        raise ValueError("No he encontrado bloques de formas de pago en el archivo.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        write_csv(records, output_path)
    else:
        write_xlsx(records, output_path)

    print(f"Archivo leido: {input_path}")
    print(f"Dias extraidos: {len(records)}")
    print(f"Salida: {output_path}")
    warnings = sorted({r["aviso_fecha"] for r in records if r["aviso_fecha"]})
    for warning in warnings:
        print(f"AVISO: {warning}", file=sys.stderr)


if __name__ == "__main__":
    main()

