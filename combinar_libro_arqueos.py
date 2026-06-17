#!/usr/bin/env python3
"""
Combina el resumen del libro de caja con el resumen de arqueos.

Uso sencillo:
  python combinar_libro_arqueos.py

Uso indicando archivos:
  python combinar_libro_arqueos.py --libro libro_de_caja20260617171340.xlsx --arqueos arqueos_1781711553.xlsx

El script comprueba que ambos archivos tengan exactamente las mismas fechas antes de generar
el Excel final. Si faltan dias en uno de los dos, para y muestra el problema.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala dependencias con: python -m pip install -r requirements.txt")

import extraer_arqueos
import extraer_libro_caja

OUTPUT_COLUMNS = [
    "Fecha",
    "Dia Semana",
    "Ingresos Tarjeta",
    "Ingresos Efectivo",
    "Salidas",
    "Retiradas Efectivo",
    "Saldo Real",
    "Saldo Teorico",
    "Descuadre",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Combina libro de caja y arqueos por fecha, comprobando que los dias coincidan."
    )
    parser.add_argument(
        "--libro",
        help="Excel de libro de caja. Si se omite, usa el libro_de_caja*.xlsx mas reciente de esta carpeta.",
    )
    parser.add_argument(
        "--arqueos",
        help="Excel/CSV de arqueos. Si se omite, usa el arqueos*.xlsx mas reciente de esta carpeta o Downloads.",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Archivo de salida .xlsx. Si se omite, crea resumen_final_caja_arqueos.xlsx.",
    )
    return parser.parse_args()


def resolve_optional_path(value: str | None, base_dir: Path) -> Path | None:
    if not value:
        return None
    path = Path(value)
    if not path.is_absolute():
        path = base_dir / path
    return path.resolve()


def extract_libro(libro_path: Path) -> list[dict[str, Any]]:
    return extraer_libro_caja.extract_cashbook(libro_path, date_mode="auto")


def extract_arqueos(arqueos_path: Path) -> list[dict[str, Any]]:
    source_rows = extraer_arqueos.read_source(arqueos_path)
    records = extraer_arqueos.group_by_close_date(source_rows)
    return [
        {
            "fecha": record["Fecha"],
            "saldo_real_arqueo": record["Saldo Real"],
            "saldo_teorico_arqueo": record["Saldo Teorico"],
            "descuadre_arqueo": record["Descuadre"],
            "retiradas_arqueo": record["Retiradas"],
        }
        for record in records
    ]


def validate_same_dates(libro_rows: list[dict[str, Any]], arqueo_rows: list[dict[str, Any]]) -> None:
    libro_dates = [str(row["fecha"]) for row in libro_rows]
    arqueo_dates = [str(row["fecha"]) for row in arqueo_rows]
    libro_set = set(libro_dates)
    arqueo_set = set(arqueo_dates)

    if libro_set == arqueo_set and libro_dates[0] == arqueo_dates[0] and libro_dates[-1] == arqueo_dates[-1]:
        return

    missing_in_arqueos = sorted(libro_set - arqueo_set)
    missing_in_libro = sorted(arqueo_set - libro_set)
    details = [
        "Las fechas del libro de caja y de arqueos no coinciden.",
        f"Libro:   {libro_dates[0] if libro_dates else 'sin datos'} -> {libro_dates[-1] if libro_dates else 'sin datos'} ({len(libro_dates)} dias)",
        f"Arqueos: {arqueo_dates[0] if arqueo_dates else 'sin datos'} -> {arqueo_dates[-1] if arqueo_dates else 'sin datos'} ({len(arqueo_dates)} dias)",
    ]
    if missing_in_arqueos:
        details.append(f"Fechas del libro que faltan en arqueos: {', '.join(missing_in_arqueos)}")
    if missing_in_libro:
        details.append(f"Fechas de arqueos que faltan en el libro: {', '.join(missing_in_libro)}")
    raise ValueError("\n".join(details))


def combine_rows(libro_rows: list[dict[str, Any]], arqueo_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    arqueos_by_date = {str(row["fecha"]): row for row in arqueo_rows}
    combined = []
    for libro_row in libro_rows:
        date = str(libro_row["fecha"])
        arqueo_row = arqueos_by_date[date]
        combined.append(
            {
                "Fecha": libro_row.get("fecha", ""),
                "Dia Semana": libro_row.get("dia_semana", ""),
                "Ingresos Tarjeta": libro_row.get("ingresos_tarjeta", 0.0),
                "Ingresos Efectivo": libro_row.get("ingresos_efectivo", 0.0),
                "Salidas": libro_row.get("salidas_caja", 0.0),
                "Retiradas Efectivo": abs(float(libro_row.get("retiradas_efectivo", 0.0) or 0.0)),
                "Saldo Real": arqueo_row.get("saldo_real_arqueo", 0.0),
                "Saldo Teorico": arqueo_row.get("saldo_teorico_arqueo", 0.0),
                "Descuadre": arqueo_row.get("descuadre_arqueo", 0.0),
            }
        )
    return combined


def write_xlsx(records: list[dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen final"
    ws.append(OUTPUT_COLUMNS)
    for record in records:
        ws.append([record.get(col, "") for col in OUTPUT_COLUMNS])

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    alt_fill = openpyxl.styles.PatternFill("solid", fgColor="F2F2F2")
    thin_side = openpyxl.styles.Side(style="thin", color="A6A6A6")
    medium_side = openpyxl.styles.Side(style="medium", color="5B5B5B")
    thin_border = openpyxl.styles.Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )
    euro_format = '#,##0.00 €'

    ws.row_dimensions[1].height = 36
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    money_columns = {
        "Ingresos Tarjeta",
        "Ingresos Efectivo",
        "Salidas",
        "Retiradas Efectivo",
        "Saldo Real",
        "Saldo Teorico",
        "Descuadre",
    }
    descuadre_col = OUTPUT_COLUMNS.index("Descuadre") + 1

    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = alt_fill
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.border = thin_border
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center",
                vertical="center",
            )
            if OUTPUT_COLUMNS[col - 1] in money_columns:
                cell.number_format = euro_format
        descuadre_cell = ws.cell(row, descuadre_col)
        if (descuadre_cell.value or 0) < 0:
            descuadre_cell.font = openpyxl.styles.Font(color="C00000")
        elif (descuadre_cell.value or 0) > 0:
            descuadre_cell.font = openpyxl.styles.Font(color="008000")

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    for col_cells in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width, 13), 24)

    # Borde exterior algo mas marcado alrededor de toda la tabla.
    min_row, max_row = 1, ws.max_row
    min_col, max_col = 1, ws.max_column
    for col in range(min_col, max_col + 1):
        top_cell = ws.cell(min_row, col)
        bottom_cell = ws.cell(max_row, col)
        top_cell.border = openpyxl.styles.Border(
            left=top_cell.border.left,
            right=top_cell.border.right,
            top=medium_side,
            bottom=top_cell.border.bottom,
        )
        bottom_cell.border = openpyxl.styles.Border(
            left=bottom_cell.border.left,
            right=bottom_cell.border.right,
            top=bottom_cell.border.top,
            bottom=medium_side,
        )
    for row in range(min_row, max_row + 1):
        left_cell = ws.cell(row, min_col)
        right_cell = ws.cell(row, max_col)
        left_cell.border = openpyxl.styles.Border(
            left=medium_side,
            right=left_cell.border.right,
            top=left_cell.border.top,
            bottom=left_cell.border.bottom,
        )
        right_cell.border = openpyxl.styles.Border(
            left=right_cell.border.left,
            right=medium_side,
            top=right_cell.border.top,
            bottom=right_cell.border.bottom,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent

    libro_path = resolve_optional_path(args.libro, base_dir) or extraer_libro_caja.resolve_input(None, base_dir)
    arqueos_path = resolve_optional_path(args.arqueos, base_dir) or extraer_arqueos.resolve_input(None, base_dir)
    output_path = resolve_optional_path(args.output, base_dir) or (base_dir / "resumen_final_caja_arqueos.xlsx")

    libro_rows = extract_libro(libro_path)
    arqueo_rows = extract_arqueos(arqueos_path)
    validate_same_dates(libro_rows, arqueo_rows)
    combined_rows = combine_rows(libro_rows, arqueo_rows)
    write_xlsx(combined_rows, output_path)

    print(f"Libro de caja: {libro_path}")
    print(f"Arqueos: {arqueos_path}")
    print(f"Dias combinados: {len(combined_rows)}")
    print(f"Salida: {output_path}")


if __name__ == "__main__":
    main()




