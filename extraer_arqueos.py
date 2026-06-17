#!/usr/bin/env python3
r"""`r`nAgrupa arqueos por fecha de cierre.

Uso sencillo:
  python extraer_arqueos.py

Uso indicando archivo:
  python extraer_arqueos.py arqueos_1781711553.xlsx
  python extraer_arqueos.py C:\Users\bafal\Downloads\arqueos_1781711553.xlsx -o resumen_arqueos.xlsx

Si no indicas archivo, busca el arqueos*.xlsx o arqueos*.csv mas reciente en esta carpeta
y despues en Downloads.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala dependencias con: python -m pip install -r requirements.txt")

OUTPUT_COLUMNS = ["Fecha", "Saldo Real", "Saldo Teorico", "Descuadre", "Retiradas"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Agrupa arqueos por fecha de cierre.")
    parser.add_argument(
        "input",
        nargs="?",
        help="Archivo arqueos .xlsx o .csv. Si se omite, busca el mas reciente.",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Salida .xlsx o .csv. Si se omite, crea resumen_arqueos_<nombre>.xlsx en esta carpeta.",
    )
    return parser.parse_args()


def normalize(text: Any) -> str:
    value = str(text or "").strip().lower()
    value = value.replace("ó", "o").replace("ò", "o")
    value = value.replace("é", "e").replace("è", "e")
    value = value.replace("á", "a").replace("à", "a")
    value = value.replace("í", "i").replace("ï", "i")
    value = value.replace("ú", "u").replace("ü", "u")
    value = value.replace("ñ", "n")
    return re.sub(r"\s+", " ", value)


def as_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)

    text = str(value).strip()
    if "," in text and "." in text:
        # Formato espanol con miles y decimales: 1.234,56
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        # Formato decimal espanol: 1234,56
        text = text.replace(",", ".")
    # Si solo hay punto, asumimos punto decimal: 1234.56

    try:
        return round(float(text), 2)
    except ValueError:
        return 0.0


def parse_datetime(value: Any) -> dt.datetime:
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"No he podido leer la fecha: {value!r}")


def candidate_files(base_dir: Path) -> list[Path]:
    downloads = Path.home() / "Downloads"
    folders = [base_dir, downloads]
    xlsx_files: list[Path] = []
    csv_files: list[Path] = []
    for folder in folders:
        if folder.exists():
            xlsx_files.extend(folder.glob("arqueos*.xlsx"))
            csv_files.extend(folder.glob("arqueos*.csv"))

    preferred = xlsx_files if xlsx_files else csv_files
    return sorted(set(preferred), key=lambda p: p.stat().st_mtime, reverse=True)


def resolve_input(input_arg: str | None, base_dir: Path) -> Path:
    if input_arg:
        path = Path(input_arg)
        if not path.is_absolute():
            path = base_dir / path
        return path.resolve()

    candidates = candidate_files(base_dir)
    if not candidates:
        raise FileNotFoundError("No encuentro ningun arqueos*.xlsx o arqueos*.csv en FBP ni en Downloads.")
    return candidates[0].resolve()


def header_index(headers: list[Any], name: str) -> int:
    wanted = normalize(name)
    normalized = [normalize(header) for header in headers]
    if wanted in normalized:
        return normalized.index(wanted)
    raise ValueError(f"No encuentro la columna {name!r}. Encabezados: {headers}")


def read_xlsx(path: Path) -> list[dict[str, Any]]:
    ws = openpyxl.load_workbook(path, data_only=True).active
    headers = [cell.value for cell in ws[1]]
    idx_fecha = header_index(headers, "Fecha cierre")
    idx_real = header_index(headers, "Saldo real")
    idx_teorico = header_index(headers, "Saldo teorico")
    idx_descuadre = header_index(headers, "Descuadre")
    idx_retirado = header_index(headers, "Importe retirado")

    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(
            {
                "fecha_cierre": values[idx_fecha],
                "saldo_real": values[idx_real],
                "saldo_teorico": values[idx_teorico],
                "descuadre": values[idx_descuadre],
                "retiradas": values[idx_retirado],
            }
        )
    return rows


def read_csv(path: Path) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with path.open("r", newline="", encoding=encoding) as fh:
                reader = csv.DictReader(fh)
                return [
                    {
                        "fecha_cierre": row.get("Fecha cierre"),
                        "saldo_real": row.get("Saldo real"),
                        "saldo_teorico": row.get("Saldo teórico") or row.get("Saldo teorico"),
                        "descuadre": row.get("Descuadre"),
                        "retiradas": row.get("Importe retirado"),
                    }
                    for row in reader
                    if any(row.values())
                ]
        except Exception as exc:
            last_error = exc
    raise ValueError(f"No he podido leer el CSV: {last_error}")


def read_source(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".csv":
        return read_csv(path)
    raise ValueError("El archivo debe ser .xlsx o .csv")


def group_by_close_date(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[dt.date, dict[str, float]] = defaultdict(
        lambda: {"Saldo Real": 0.0, "Saldo Teorico": 0.0, "Descuadre": 0.0, "Retiradas": 0.0}
    )
    for row in rows:
        close_date = parse_datetime(row["fecha_cierre"]).date()
        grouped[close_date]["Saldo Real"] += as_number(row["saldo_real"])
        grouped[close_date]["Saldo Teorico"] += as_number(row["saldo_teorico"])
        grouped[close_date]["Descuadre"] += as_number(row["descuadre"])
        grouped[close_date]["Retiradas"] += as_number(row["retiradas"])

    records = []
    for date in sorted(grouped):
        item = grouped[date]
        records.append(
            {
                "Fecha": date.isoformat(),
                "Saldo Real": round(item["Saldo Real"], 2),
                "Saldo Teorico": round(item["Saldo Teorico"], 2),
                "Descuadre": round(item["Descuadre"], 2),
                "Retiradas": round(item["Retiradas"], 2),
            }
        )
    return records


def write_xlsx(records: list[dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen arqueos"
    ws.append(OUTPUT_COLUMNS)
    for record in records:
        ws.append([record[col] for col in OUTPUT_COLUMNS])

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in range(2, ws.max_row + 1):
        for col in range(2, 6):
            ws.cell(row, col).number_format = '#,##0.00'

    widths = {"A": 14, "B": 16, "C": 18, "D": 14, "E": 14}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    wb.save(output_path)


def format_decimal(value: Any) -> Any:
    if isinstance(value, float):
        return f"{value:.2f}".replace(".", ",")
    return value


def write_csv(records: list[dict[str, Any]], output_path: Path) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS, delimiter=";")
        writer.writeheader()
        for record in records:
            writer.writerow({col: format_decimal(record[col]) for col in OUTPUT_COLUMNS})


def default_output(input_path: Path, base_dir: Path) -> Path:
    return base_dir / f"resumen_arqueos_{input_path.stem}.xlsx"


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    input_path = resolve_input(args.input, base_dir)
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_path = Path(args.output).resolve() if args.output else default_output(input_path, base_dir)
    rows = read_source(input_path)
    records = group_by_close_date(rows)
    if not records:
        raise ValueError("No he encontrado arqueos para agrupar.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        write_csv(records, output_path)
    else:
        write_xlsx(records, output_path)

    print(f"Archivo leido: {input_path}")
    print(f"Dias agrupados: {len(records)}")
    print(f"Salida: {output_path}")


if __name__ == "__main__":
    main()



